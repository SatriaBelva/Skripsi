"""
=============================================================
  ANALISIS KOLABORASI INTERNASIONAL - UNIVERSITAS JEMBER
  Input  : file CSV export dari Scopus
  Output : hasil_per_artikel.xlsx (satu baris per artikel)
=============================================================
Cara pakai:
  1. Letakkan script ini & file CSV Scopus dalam 1 folder
  2. Ubah INPUT_FILE di bawah sesuai nama file CSV kamu
  3. Jalankan: python analisis_kolaborasi_scopus.py
=============================================================
"""

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from collections import Counter
import os, sys

# ─────────────────────────────────────────────────────────
# KONFIGURASI ← ubah sesuai kebutuhan
# ─────────────────────────────────────────────────────────
INPUT_FILE  = 'scopus_export_POLIJE.csv'

KAMPUS_AKTIF = 'POLIJE' 

# Mapping kata kunci spesifik agar data ke-3 kampus tidak tercampur
KEYWORDS_MAP = {
    'UNEJ': ['universitas jember', 'university of jember', 'jember university', 'unej'],
    'UNMUH': ['muhammadiyah jember', 'muhammadiyah university of jember', 'unmuh'],
    'POLIJE': ['politeknik negeri jember', 'state polytechnic of jember', 'polije']
}
TARGET_KEYWORDS = KEYWORDS_MAP[KAMPUS_AKTIF]
INDO_KEYWORD  = 'indonesia'
OUTPUT_FILE = f'Hasil Artikel {KAMPUS_AKTIF}.xlsx'

# ─────────────────────────────────────────────────────────
# FUNGSI PEMBANTU
# ─────────────────────────────────────────────────────────

def is_target_campus(text):
    t = (text or '').lower()
    return any(k in t for k in TARGET_KEYWORDS)

def is_indonesia(text):
    return INDO_KEYWORD in (text or '').lower()

def extract_country(affil):
    parts = [p.strip() for p in (affil or '').split(',')]
    for p in reversed(parts):
        if p and len(p) > 1 and not p.isdigit():
            return p
    return 'Unknown'

def parse_authors_affil(raw):
    """Parse kolom 'Authors with affiliations' -> list of (author, affil)."""
    if not isinstance(raw, str) or not raw.strip():
        return []
    result = []
    for entry in raw.split(';'):
        entry = entry.strip()
        if not entry:
            continue
        idx = entry.find(',')
        if idx > 0:
            result.append((entry[:idx].strip(), entry[idx+1:].strip()))
        else:
            result.append((entry, ''))
    return result

def get_intl_affiliations(affiliations_col):
    """Ambil afiliasi non-Indonesia dari kolom Affiliations."""
    if not isinstance(affiliations_col, str):
        return []
    return [
        a.strip() for a in affiliations_col.split(';')
        if a.strip() and not is_indonesia(a)
    ]


# ─────────────────────────────────────────────────────────
# PROSES DATA
# ─────────────────────────────────────────────────────────

def process(df):
    rows = []
    for i, row in df.iterrows():
        title  = str(row.get('Title',        f'[No Title {i+1}]'))
        year   = row.get('Year',             '')
        doi    = row.get('DOI',              '')
        cited  = row.get('Cited by',         0)
        source = row.get('Source title',     '')

        entries   = parse_authors_affil(row.get('Authors with affiliations', ''))
        internal_list = [] # Menggantikan unej_list
        intl_list = []

        for author, affil in entries:
            if is_target_campus(affil): # Menggunakan fungsi yang baru
                internal_list.append(author)
            elif not is_indonesia(affil) and affil:
                country = extract_country(affil)
                inst    = affil.split(',')[0].strip()
                intl_list.append(f"{author} [{inst}, {country}]")

        # Negara mitra dari kolom Affiliations (lebih bersih)
        intl_countries = set()
        for aff in get_intl_affiliations(row.get('Affiliations', '')):
            c = extract_country(aff)
            if c and c != 'Unknown':
                intl_countries.add(c)

        rows.append({
            'No'                     : i + 1,
            'Judul'                  : title,
            'Tahun'                  : year,
            'Source'                 : source,
            'DOI'                    : doi,
            'Cited By'               : cited,
            'Author UNEJ'            : ' | '.join(internal_list),
            'Co-Author Internasional': ' | '.join(intl_list),
            'Negara Mitra'           : ' | '.join(sorted(intl_countries)),
            'Jumlah Negara'          : len(intl_countries),
        })

    return rows


# ─────────────────────────────────────────────────────────
# BUAT FILE EXCEL DENGAN FORMATTING
# ─────────────────────────────────────────────────────────

def save_excel(rows, output_file):
    wb = Workbook()
    ws = wb.active
    ws.title = 'Hasil Per Artikel'

    # Warna
    HEADER_BG  = '1F4E79'  # biru tua
    HEADER_FG  = 'FFFFFF'  # putih
    ROW_ODD    = 'FFFFFF'  # baris ganjil - putih
    ROW_EVEN   = 'EAF2FB'  # baris genap - biru muda
    ACCENT_ODD  = 'EEF5FC' # kolom analisis baris ganjil
    ACCENT_EVEN = 'D9E8F5' # kolom analisis baris genap

    thin   = Side(style='thin', color='D0D0D0')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    headers = [
        'No', 'Judul', 'Tahun', 'Source', 'DOI',
        'Cited By', 'Author Asli', 'Co-Author Internasional',
        'Negara Mitra', 'Jumlah Negara'
    ]

    # Lebar kolom
    col_widths = [5, 55, 7, 30, 35, 10, 30, 50, 35, 14]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # Tulis header
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font      = Font(name='Arial', bold=True, color=HEADER_FG, size=10)
        cell.fill      = PatternFill('solid', start_color=HEADER_BG)
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border    = border
    ws.row_dimensions[1].height = 32

    # Tulis data
    for row_idx, r in enumerate(rows, 2):
        is_even = (row_idx % 2 == 0)

        values = [
            r['No'], r['Judul'], r['Tahun'], r['Source'], r['DOI'],
            r['Cited By'], r['Author UNEJ'], r['Co-Author Internasional'],
            r['Negara Mitra'], r['Jumlah Negara'],
        ]

        for col_idx, value in enumerate(values, 1):
            cell        = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font   = Font(name='Arial', size=9)
            cell.border = border

            # Kolom numerik: rata tengah
            if col_idx in (1, 3, 6, 10):
                cell.alignment = Alignment(horizontal='center', vertical='top')
            else:
                cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)

            # Warna latar: kolom analisis utama lebih gelap sedikit
            if col_idx in (7, 8, 9, 10):
                bg = ACCENT_EVEN if is_even else ACCENT_ODD
            else:
                bg = ROW_EVEN if is_even else ROW_ODD

            cell.fill = PatternFill('solid', start_color=bg)

        ws.row_dimensions[row_idx].height = 40

    # Freeze baris header & aktifkan auto-filter
    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = f'A1:{get_column_letter(len(headers))}1'

    wb.save(output_file)


# ─────────────────────────────────────────────────────────
# MAIN
# ─────────────────────────────────────────────────────────

def main():
    if not os.path.exists(INPUT_FILE):
        print(f'\n ERROR: File "{INPUT_FILE}" tidak ditemukan.')
        print(f'   Pastikan file CSV Scopus ada di folder yang sama dengan script ini.')
        print(f'   Folder saat ini: {os.getcwd()}')
        sys.exit(1)

    print(f'\n Membaca: {INPUT_FILE}')
    df = pd.read_csv(INPUT_FILE, encoding='utf-8-sig')
    print(f' Total dokumen: {len(df):,}')

    for col in ['Authors with affiliations', 'Affiliations', 'Title']:
        if col not in df.columns:
            print(f' PERINGATAN: Kolom "{col}" tidak ditemukan — pastikan di-export dari Scopus.')

    print(' Memproses data...')
    rows = process(df)

    print(' Menyimpan Excel...')
    save_excel(rows, OUTPUT_FILE)

    print(f'\n Selesai! File tersimpan -> {OUTPUT_FILE}')
    print(f'   Total baris data: {len(rows):,}')

    # Ringkasan negara di terminal
    all_countries = []
    for r in rows:
        if r['Negara Mitra']:
            all_countries.extend(r['Negara Mitra'].split(' | '))
    ctr = Counter(all_countries)

    print(f'\n Top 10 Negara Mitra:')
    print(f'   {"Negara":<30} {"Dok":>5}')
    print(f'   {"-"*35}')
    for country, count in ctr.most_common(10):
        print(f'   {country:<30} {count:>5}')
    print()


if __name__ == '__main__':
    main()
